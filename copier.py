import openpyxl
import random

for num in range(1, 101):
    path = f"Performance_Metrics/site_{num}.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    for i in range(1, 27):
        for j in range(1, 8):
            cell_obj = sheet_obj.cell(row=i, column=j)
            c1 = sheet.cell(row=i, column=j)
            c1.value = cell_obj.value

    for i in range(2, 27):
        random_row = random.randint(2, 26)
        for j in range(2, 8):
            cell_obj = sheet_obj.cell(row=random_row, column=j)
            c1 = sheet.cell(row=i+25, column=j)
            c1.value = cell_obj.value

    wb.save(f"site_{num}.xlsx")
