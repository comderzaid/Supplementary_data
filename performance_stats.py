import numpy
import openpyxl
import statistics as st


wb = openpyxl.Workbook()
sheet = wb.active


for i in range(1, 101):
    path = f"site_{i}.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for j in range(2, 8):
        values = []
        for k in range(2, 52):
            cell_obj = sheet_obj.cell(row=k, column=j)
            try:
                value = float(cell_obj.value)
            except:
                value = 4.2
            values.append(value)

        mean = numpy.mean(values)
        c2 = sheet.cell(row=i, column=j-1)
        c2.value = mean
wb.save("Performance.xlsx")
