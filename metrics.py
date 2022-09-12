import requests
from bs4 import BeautifulSoup
import openpyxl

# path = "Webby_2017_awards.xlsx"
# wb = openpyxl.Workbook()
# sheet = wb.active
# wb_obj = openpyxl.load_workbook(path)
# sheet_obj = wb_obj.active

# for j in range(0, 100):
#     cell_obj = sheet_obj.cell(row=j+2, column=1)
#     print(cell_obj.value)
#     try:
#         response = requests.get(cell_obj.value)
#     except:
#         response = requests.get("https://www.apple.com")

#     if response.status_code != 200:
#         response = requests.get("https://www.apple.com")

#     soup = BeautifulSoup(response.text, "html.parser")
#     links = len(soup.find_all('a', href=True)) + \
#         len(soup.find_all('link', href=True))  # number of links

#     body_words = len(soup.find("body").text)   # body text words

#     lists = len(soup.find_all("ol")) + \
#         len(soup.find_all("ul"))  # number of lists

#     tables = len(soup.find_all("table"))  # number of tabeles

#     if soup.find("title"):
#         title_length = len(soup.find("title").text)  # page title length
#     else:
#         title_length = 6

#     page_size = 100 * body_words  # page size

#     graphics = len(soup.find_all("img")) + \
#         len(soup.find_all("svg")) + \
#         len(soup.find_all("canvas"))  # Number of graphics

#     text_emphasis = len(soup.find_all("b")) + \
#         len(soup.find_all("strong")) + len(soup.find_all("i")) + \
#         len(soup.find_all("em")) + len(soup.find_all("u")) + \
#         len(soup.find_all("del")) + \
#         len(soup.find_all("s")) + \
#         len(soup.find_all("sub"))  # Body Text Emphasis

#     number_of_ = soup.find("body").text.count("!")  # number of !

#     script_tag = len(soup.find_all("script"))  # Number of Script

#     embedded_links = len(soup.find_all('a', href=True))  # Embedded links

#     redirecting_links = len(soup.find_all(
#         'a', href=True))  # Redirecting Links

#     in_page_link = int(
#         len(soup.find_all('link', href=True))/10)  # In Page links

#     frame_tags = len(soup.find_all("frame"))  # Frame Tags

#     total_number_of_words = len(soup.text)  # Total Number of Words

#     meta_tags = len(soup.find_all("meta"))  # Number of Meta Tags

#     metrics = [links, body_words, lists, tables,
#                title_length, page_size, graphics, text_emphasis, number_of_, script_tag, embedded_links, redirecting_links, in_page_link, frame_tags, total_number_of_words, meta_tags]

#     print(f"Success - {j+2}")
#     for i in range(0, len(metrics)):
#         c2 = sheet.cell(row=j+1, column=i+1)
#         c2.value = metrics[i]
#     wb.save("metric_data_2017.xlsx")


response = requests.get("https://nhentai.net/")
soup = BeautifulSoup(response.text, "html.parser")

metrics_format = """
1. Number of Links :- {number_of_link}
2. Body Text Words :- {body_text_words}
3. Number of list :- {number_of_list}
4. Number of tables :- {number_of_tables}
5. Page title length :- {title_length}
6. Page size :- {page_size}
7. Number of graphics :- {graphics}
8. Text emphasis :- {text_emphasis}
9. Number of !’s :- {number_of_exclam}
10. Number of script:- {number_of_script}
11. Embedded links :- {embedded_links}
12. Number of redirecting links :- {redirecting_links}
13. Number of in-page links :- {in_page_link}
14. Frames :- {frames}
15. Total Number of words :- {number_of_words}
16. Number of meta tags :- {number_of_meta_tags}
"""

links = len(soup.find_all('a', href=True)) + \
    len(soup.find_all('link', href=True))  # number of links

body_words = len(soup.find("body").text)   # body text words

lists = len(soup.find_all("ol")) + len(soup.find_all("ul"))  # number of lists

tables = len(soup.find_all("table"))  # number of tabeles

title_length = len(soup.find("title").text)  # page title length

page_size = 100 * body_words  # page size

graphics = len(soup.find_all("img")) + \
    len(soup.find_all("svg")) + \
    len(soup.find_all("canvas"))  # Number of graphics

text_emphasis = len(soup.find_all("b")) + \
    len(soup.find_all("strong")) + len(soup.find_all("i")) + \
    len(soup.find_all("em")) + len(soup.find_all("u")) + \
    len(soup.find_all("del")) + \
    len(soup.find_all("s")) + len(soup.find_all("sub"))  # Body Text Emphasis

number_of_ = soup.find("body").text.count("!")  # number of !

script_tag = len(soup.find_all("script"))  # Number of Script

embedded_links = len(soup.find_all('a', href=True))  # Embedded links

redirecting_links = len(soup.find_all('a', href=True))  # Redirecting Links

in_page_link = int(len(soup.find_all('link', href=True))/10)  # In Page links

frame_tags = len(soup.find_all("frame"))  # Frame Tags

total_number_of_words = len(soup.text)  # Total Number of Words

meta_tags = len(soup.find_all("meta"))  # Number of Meta Tags

print(metrics_format.format(
    number_of_link=links,
    body_text_words=body_words,
    number_of_list=lists,
    number_of_tables=tables,
    title_length=title_length,
    page_size=page_size,
    graphics=graphics,
    text_emphasis=text_emphasis,
    number_of_exclam=number_of_,
    number_of_script=script_tag,
    embedded_links=embedded_links,
    redirecting_links=redirecting_links,
    in_page_link=in_page_link,
    frames=frame_tags,
    number_of_words=total_number_of_words,
    number_of_meta_tags=meta_tags
))
