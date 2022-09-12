import numpy
import openpyxl
import statistics as st
from time import time
from datetime import datetime


def stat(year: int):
    start = time()
    path = f"Webby_{year}_awards.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    values = []

    for i in range(2, 18):
        for j in range(2, 102):
            cell_obj = sheet_obj.cell(row=j, column=i)
            values.append(int(cell_obj.value))
        temp = numpy.array(values)
        stats = [min(values), max(values), numpy.mean(values),
                 numpy.median(values), st.mode(temp), numpy.std(values)]
        for k in range(0, len(stats)):
            c2 = sheet.cell(row=i, column=k+2)
            c2.value = stats[k]
        wb.save(f"Descriptive_statistics_{year}.xlsx")
        values = []

    end = time()
    takenTime = datetime.utcfromtimestamp(end - start)
    takenTime = f"""{f"{takenTime.minute} Min(s)" if takenTime.minute > 0 else ""}{f"{takenTime.second} Sec(s)" if takenTime.second > 0 else "0.0 Sec(s)"}"""
    print(f"Success \nTime Taken :- {takenTime}")


year = input("Enter year whose stats is to be calculated:- ")
stat(int(year))
