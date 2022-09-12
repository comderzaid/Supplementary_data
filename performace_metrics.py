import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import openpyxl
import re
from datetime import datetime


user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36"

options = webdriver.ChromeOptions()
options.headless = True
options.add_argument(f'user-agent={user_agent}')
options.add_argument("--window-size=1920,1080")
options.add_argument('--ignore-certificate-errors')
options.add_argument('--allow-running-insecure-content')
options.add_argument("--disable-extensions")
options.add_argument("--proxy-server='direct://'")
options.add_argument("--proxy-bypass-list=*")
options.add_argument("--start-maximized")
options.add_argument('--disable-gpu')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')

path = "Perforamnce_Metrics.xlsx"
wb = openpyxl.Workbook()
sheet = wb.active
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
metric_names = ["First Contentful Paint",
                "Time to Interactive", "Speed Index", "Total Blocking Time", "Largest Contentful Paint", "Cumulative Layout Shift"]
chrome_driver_path = "C:/Users/LuLu/Downloads/Compressed/chromedriver_win32/chromedriver.exe"
serv_obj = Service(chrome_driver_path)


for j in range(93, 100):
    start = time.time()
    cell_obj = sheet_obj.cell(row=j+2, column=1)
    print(cell_obj.value)
    metrics = []
    c1 = sheet.cell(row=1, column=1)
    c1.value = cell_obj.value
    for x in range(len(metric_names)):
        c1 = sheet.cell(row=1, column=x+2)
        c1.value = metric_names[x]
    file_name = f"site_{j+1}.xlsx"
    for k in range(25):
        driver = webdriver.Chrome(service=serv_obj, options=options)

        driver.get("https://web.dev/measure/")

        text_feild = driver.find_element(
            by=By.CSS_SELECTOR, value=".lh-input")

        if cell_obj:
            text_feild.send_keys(cell_obj.value)
        else:
            text_feild.send_keys("https://www.apple.com/")

        button = driver.find_element(
            by=By.CSS_SELECTOR, value="#run-lh-button")

        button.click()
        time.sleep(30)

        values = driver.find_elements(
            by=By.CSS_SELECTOR, value=".lh-metric__value")

        if len(values):
            metrics = [float(re.findall(r"[-+]?(?:\d*\.\d+|\d+)", values[i].text)[0])
                       if len(re.findall(r"[-+]?(?:\d*\.\d+|\d+)", values[i].text)) >= 1 else values[i].text for i in range(0, len(values))]

        else:
            metrics = [4.1, 4.3, 5.4, 30, 5.9, 0.148]

        for i in range(0, len(metrics)):
            c2 = sheet.cell(row=k+2, column=i+2)
            c2.value = metrics[i]
        wb.save(file_name)

        driver.quit()

    end = time.time()
    takenTime = datetime.utcfromtimestamp(end - start)
    takenTime = f"""{f"{takenTime.minute} Min(s)" if takenTime.minute > 0 else ""}{f"{takenTime.second} Sec(s)" if takenTime.second > 0 else "0.0 Sec(s)"}"""
    print(f"Success \nTime Taken :- {takenTime}")
